# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import openpyxl
import psycopg2

from config import host, user, password, db_name

def Value_or_NULL(value):   #Выдать NULL, если в таблице это значение помечено как NULL
    if str(value).lower() == 'null':
        return 'null'
    else:
        return value

def checkTable(dbcon, tablename):  #Это было стырено
    dbcur = dbcon.cursor()
    dbcur.execute(
        """
        SELECT COUNT(*)
        FROM information_schema.tables
        WHERE table_name = '{0}'
        """.format(tablename.replace('\'', '\'\'')))
    if dbcur.fetchone()[0] == 1:
        dbcur.close()
        return True

    dbcur.close()
    return False

def open_table():
    pass

try:
    connection = psycopg2.connect(
        host=host,
        user=user,
        password=password,
        database=db_name
    )
    connection.autocommit = True
    # str(file_table[i][j].value).replace(' 00:00:00', '')
    # with connection.cursor() as cursor:
    #     cursor.execute(
    #         'SELECT version();'
    #     )
    #     print(f"Серверок у нас: {cursor.fetchone()}")

    name_file_1 = 'C:/Users/cheog/Desktop/таблица_после_ереси'#input('Введите адрес и название архива: ') #  Ввод данных из экселя
    try: #              Двумерный массив с данными. Нулевая строка содержит названия столбцов
        file_table = openpyxl.open(f'{name_file_1}.xlsx', read_only=True).active
        table_list_1 = [[str(file_table[i][j].value) for j in range(0, file_table.max_column)] for i in
                      range(1, file_table.max_row + 1)]
        for row in table_list_1:
            if ' 00:00:00' in row[2]:
                row[2] = row[2].replace(' 00:00:00', '')
                row[2] = row[2].split('-')[2] + '-' + row[2].split('-')[1] + '-' + row[2].split('-')[0]
    except Exception as ex:
        print('Такой файл не найден', ex)

    name_file_2 = 'C:/Users/cheog/Desktop/таблица_до_ереси'  # input('Введите адрес и название архива: ') #  Ввод данных из экселя
    try:  # Двумерный массив с данными. Нулевая строка содержит названия столбцов
        file_table = openpyxl.open(f'{name_file_2}.xlsx', read_only=True).active
        table_list_2 = [[str(file_table[i][j].value) for j in range(0, file_table.max_column)] for i in
                      range(1, file_table.max_row + 1)]
        # for rows in table_list:
        #     print(rows)
    except Exception as ex:
        print('Такой файл не найден', ex)

    with connection.cursor() as cursor: # Создание таблицы "ПОСЛЕ ереси"
        name_table_1 = 'after_the_heresy'#input('Введите название таблы')
        if checkTable(connection, name_table_1) != True:
            cursor.execute(
                f"""
                CREATE TABLE {name_table_1}(
                name_order VARCHAR(40),
                name_primarch VARCHAR(40),
                date_of_found DATE,
                loyalty boolean,
                number INT PRIMARY KEY,
                home_world VARCHAR(40)
                );
                """
            )
        else:
            print(f'Таблица {name_table_1} найдена')


    with connection.cursor() as cursor: # Создание таблицы "ДО ереси"
        name_table_2 = 'before_the_heresy'#input('Введите название таблы')
        if checkTable(connection, name_table_2) != True:
            cursor.execute(
                f"""
                CREATE TABLE {name_table_2}(
                name_order VARCHAR(40),
                number INT PRIMARY KEY
                );
                """
            )
        else:
            print(f'Таблица {name_table_2} найдена')


    with connection.cursor() as cursor: #Заполнение таблицы ПОСЛЕ ереси
        for row in table_list_1[1:]:
            cursor.execute(
                f"""
                INSERT INTO after_the_heresy(name_order, name_primarch, date_of_found, loyalty, number, home_world) VALUES
                ('{row[0]}',
                 '{row[1]}',
                 {f"TO_DATE('{row[2]}', 'dd.mm.yyyy')" if row[2] != 'null' else 'null::timestamp'},
                 {str(bool(int(row[3]))) if str(row[3]).lower() != 'null' else 'null::boolean'},
                 '{row[4]}',
                 '{row[5]}'
                 );
                """
            )
        print('У нас пополнение')

    with connection.cursor() as cursor: #Заполнение таблицы ДО ереси
        for row in table_list_2[1:]:
            cursor.execute(
                f"""
                INSERT INTO after_the_heresy(name_order, number) VALUES
                ('{row[0]}',
                 '{row[1]}'
                 );
                """
            )
        print('У нас пополнение')

    # with connection.cursor() as cursor:
    #     cursor.execute(
    #         """DROP TABLE after_the_heresy"""
    #     )

except Exception as ex:
    print('Лох, потому что: ', ex)
finally:
    if connection:
        connection.close()
        print('Форточка закрыта')

#test = openpyxl.open('C:/Users/cheog/Desktop/ФИЛЬМЫ.xlsx', 'rt')
#print(test.active['B2'].value)

# with open('C:/Users/cheog/Desktop/ФИЛЬМЫ.xlsx', 'rt') as inp: #, encoding = 'utf-8'
#    print(inp.read())

